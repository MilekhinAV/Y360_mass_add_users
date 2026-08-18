#!/usr/bin/env python3
"""Массовое создание сотрудников и иерархии подразделений в Яндекс 360 из XLSX."""

from __future__ import annotations

import argparse
import csv
import getpass
import json
import logging
import os
import random
import re
import socket
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - проверяется при запуске без зависимости
    print(
        "Ошибка: не установлен openpyxl. Выполните: "
        "python3 -m pip install -r requirements.txt",
        file=sys.stderr,
    )
    raise SystemExit(2)


API_BASE_URL = "https://api360.yandex.net"
ROOT_DEPARTMENT_ID = 1
FIXED_HEADERS = [
    "last_name",
    "first_name",
    "middle_name",
    "login",
    "password",
    "must_change_password",
    "position",
    "gender",
    "birthday",
    "language",
    "work_phone",
    "mobile_phone",
    "personal_email",
]
RETRYABLE_HTTP_CODES = {429, 500, 502, 503, 504}
EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")
LOGGER = logging.getLogger("yandex360-import")


class ImportValidationError(Exception):
    """Ошибка входных данных, при которой отправлять запросы на изменение нельзя."""


class ApiError(Exception):
    """Ошибка запроса к API Яндекс 360."""

    def __init__(
        self,
        message: str,
        *,
        status: int | None = None,
        request_id: str = "",
        response_body: str = "",
    ) -> None:
        super().__init__(message)
        self.status = status
        self.request_id = request_id
        self.response_body = response_body


@dataclass(frozen=True)
class UserRow:
    row_number: int
    last_name: str
    first_name: str
    middle_name: str
    login: str
    password: str
    must_change_password: bool | None
    position: str
    gender: str
    birthday: str
    language: str
    work_phone: str
    mobile_phone: str
    personal_email: str
    department_path: tuple[str, ...]


@dataclass
class ReportRow:
    row_number: int
    login: str
    full_name: str
    department_path: str
    status: str
    user_id: str = ""
    message: str = ""
    request_id: str = ""


@dataclass
class Summary:
    departments_existing: int = 0
    departments_created: int = 0
    departments_planned: int = 0
    users_created: int = 0
    users_planned: int = 0
    users_failed: int = 0
    users_skipped: int = 0
    reports: list[ReportRow] = field(default_factory=list)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Создаёт подразделения и сотрудников Яндекс 360 по данным листа "
            "'Шаблон' в XLSX."
        )
    )
    parser.add_argument("xlsx", type=Path, help="Путь к XLSX-файлу")
    parser.add_argument(
        "--sheet",
        default="Шаблон",
        help="Имя листа с пользователями (по умолчанию: Шаблон)",
    )
    parser.add_argument("--dry-run", action="store_true", help="Проверить и показать план без POST-запросов")
    parser.add_argument("--token", help="OAuth-токен; безопаснее использовать OAUTH_TOKEN или интерактивный ввод")
    parser.add_argument("--org-id", help="ID организации; также читается из ORG_ID")
    parser.add_argument(
        "--env-file",
        type=Path,
        default=Path(".env"),
        help="Файл с OAUTH_TOKEN и ORG_ID (по умолчанию: .env)",
    )
    parser.add_argument(
        "--request-delay",
        type=float,
        default=0.25,
        help="Минимальная пауза между API-запросами, секунд (по умолчанию: 0.25)",
    )
    parser.add_argument(
        "--max-retries",
        type=int,
        default=5,
        help="Число попыток для 429/5xx и сетевых ошибок (по умолчанию: 5)",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=30.0,
        help="Тайм-аут одного HTTP-запроса, секунд (по умолчанию: 30)",
    )
    parser.add_argument(
        "--report",
        type=Path,
        help="Путь итогового CSV-отчёта; по умолчанию создаётся рядом с XLSX",
    )
    parser.add_argument("--verbose", action="store_true", help="Подробный журнал")
    args = parser.parse_args(argv)
    if args.request_delay < 0:
        parser.error("--request-delay не может быть отрицательным")
    if args.max_retries < 1:
        parser.error("--max-retries должен быть не меньше 1")
    if args.timeout <= 0:
        parser.error("--timeout должен быть больше 0")
    return args


def configure_logging(verbose: bool) -> None:
    logging.basicConfig(
        level=logging.DEBUG if verbose else logging.INFO,
        format="%(asctime)s %(levelname)s: %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def load_env_file(path: Path) -> dict[str, str]:
    """Читает простой .env без внешней зависимости python-dotenv."""
    if not path.is_file():
        return {}
    result: dict[str, str] = {}
    try:
        text = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ImportValidationError(f"Файл {path} должен быть в UTF-8: {exc}") from exc
    for line_number, raw_line in enumerate(text.splitlines(), start=1):
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("export "):
            line = line[7:].strip()
        if "=" not in line:
            raise ImportValidationError(f"{path}, строка {line_number}: ожидается ИМЯ=значение")
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if len(value) >= 2 and value[0] == value[-1] and value[0] in {'"', "'"}:
            value = value[1:-1]
        result[key] = value
    return result


def resolve_credentials(args: argparse.Namespace) -> tuple[str, int]:
    env_file = load_env_file(args.env_file)
    token = (args.token or os.environ.get("OAUTH_TOKEN") or env_file.get("OAUTH_TOKEN") or "").strip()
    org_raw = (args.org_id or os.environ.get("ORG_ID") or env_file.get("ORG_ID") or "").strip()
    if not token:
        if sys.stdin.isatty():
            token = getpass.getpass("OAuth-токен: ").strip()
        else:
            raise ImportValidationError("Не указан OAuth-токен: OAUTH_TOKEN, --token или .env")
    if not org_raw:
        if sys.stdin.isatty():
            org_raw = input("ID организации: ").strip()
        else:
            raise ImportValidationError("Не указан ID организации: ORG_ID, --org-id или .env")
    if not org_raw.isdigit() or int(org_raw) <= 0:
        raise ImportValidationError("ORG_ID должен быть положительным целым числом")
    return token, int(org_raw)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_bool(value: Any, row_number: int) -> bool | None:
    if value is None or clean_text(value) == "":
        return None
    if isinstance(value, bool):
        return value
    normalized = clean_text(value).casefold()
    true_values = {"true", "1", "да", "yes", "y"}
    false_values = {"false", "0", "нет", "no", "n"}
    if normalized in true_values:
        return True
    if normalized in false_values:
        return False
    raise ImportValidationError(
        f"Строка {row_number}: must_change_password должен быть true/false, 1/0 или да/нет"
    )


def parse_birthday(value: Any, row_number: int) -> str:
    if value is None or clean_text(value) == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    try:
        return datetime.strptime(text, "%Y-%m-%d").date().isoformat()
    except ValueError as exc:
        raise ImportValidationError(
            f"Строка {row_number}: birthday должен иметь формат YYYY-MM-DD"
        ) from exc


def is_blank_row(values: Iterable[Any]) -> bool:
    return all(clean_text(value) == "" for value in values)


def read_users_from_xlsx(path: Path, sheet_name: str) -> list[UserRow]:
    if not path.is_file():
        raise ImportValidationError(f"XLSX-файл не найден: {path}")
    if path.suffix.casefold() != ".xlsx":
        raise ImportValidationError("Поддерживается только формат .xlsx")
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ImportValidationError(f"Не удалось открыть XLSX: {exc}") from exc
    try:
        if sheet_name not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise ImportValidationError(
                f"Лист '{sheet_name}' не найден. Доступные листы: {available}"
            )
        sheet = workbook[sheet_name]
        header_values = [clean_text(cell.value) for cell in sheet[1]]
        if len(header_values) < len(FIXED_HEADERS):
            raise ImportValidationError(
                f"На листе '{sheet_name}' должно быть не меньше 13 колонок A–M"
            )
        actual_fixed = header_values[: len(FIXED_HEADERS)]
        if actual_fixed != FIXED_HEADERS:
            mismatches = [
                f"{index + 1}: ожидалось '{expected}', найдено '{actual}'"
                for index, (expected, actual) in enumerate(zip(FIXED_HEADERS, actual_fixed))
                if expected != actual
            ]
            raise ImportValidationError("Некорректные заголовки A–M: " + "; ".join(mismatches))

        department_headers = header_values[len(FIXED_HEADERS) :]
        last_department_column = -1
        for index, header in enumerate(department_headers):
            if header:
                last_department_column = index
        if last_department_column >= 0:
            used_department_headers = department_headers[: last_department_column + 1]
            if any(not header for header in used_department_headers):
                raise ImportValidationError("Заголовки уровней подразделений N+ не должны содержать пропуски")
        else:
            used_department_headers = []

        rows: list[UserRow] = []
        seen_logins: dict[str, int] = {}
        max_columns = len(FIXED_HEADERS) + len(used_department_headers)
        for row_number, raw_cells in enumerate(
            sheet.iter_rows(min_row=2, max_col=max_columns, values_only=True), start=2
        ):
            values = list(raw_cells)
            if is_blank_row(values):
                continue
            fixed = values[: len(FIXED_HEADERS)]
            departments = [clean_text(value) for value in values[len(FIXED_HEADERS) :]]
            seen_gap = False
            for level, department in enumerate(departments, start=1):
                if not department:
                    seen_gap = True
                elif seen_gap:
                    raise ImportValidationError(
                        f"Строка {row_number}: перед уровнем подразделения {level} есть пустая ячейка"
                    )
            department_path = tuple(value for value in departments if value)

            last_name = clean_text(fixed[0])
            first_name = clean_text(fixed[1])
            login = clean_text(fixed[3])
            missing = [
                name
                for name, value in (
                    ("last_name", last_name),
                    ("first_name", first_name),
                    ("login", login),
                )
                if not value
            ]
            if missing:
                raise ImportValidationError(
                    f"Строка {row_number}: не заполнены обязательные поля: {', '.join(missing)}"
                )
            login_key = login.casefold()
            if login_key in seen_logins:
                raise ImportValidationError(
                    f"Строка {row_number}: логин '{login}' уже указан в строке {seen_logins[login_key]}"
                )
            seen_logins[login_key] = row_number

            personal_email = clean_text(fixed[12])
            if personal_email and not EMAIL_RE.fullmatch(personal_email):
                raise ImportValidationError(
                    f"Строка {row_number}: некорректный personal_email '{personal_email}'"
                )
            rows.append(
                UserRow(
                    row_number=row_number,
                    last_name=last_name,
                    first_name=first_name,
                    middle_name=clean_text(fixed[2]),
                    login=login,
                    password=clean_text(fixed[4]),
                    must_change_password=parse_bool(fixed[5], row_number),
                    position=clean_text(fixed[6]),
                    gender=clean_text(fixed[7]),
                    birthday=parse_birthday(fixed[8], row_number),
                    language=clean_text(fixed[9]),
                    work_phone=clean_text(fixed[10]),
                    mobile_phone=clean_text(fixed[11]),
                    personal_email=personal_email,
                    department_path=department_path,
                )
            )
        if not rows:
            raise ImportValidationError(f"На листе '{sheet_name}' нет строк с пользователями")
        return rows
    finally:
        workbook.close()


class ApiClient:
    def __init__(
        self,
        token: str,
        org_id: int,
        *,
        request_delay: float,
        max_retries: int,
        timeout: float,
        base_url: str = API_BASE_URL,
    ) -> None:
        self.token = token
        self.org_id = org_id
        self.request_delay = request_delay
        self.max_retries = max_retries
        self.timeout = timeout
        self.base_url = base_url.rstrip("/")
        self._last_request_at = 0.0

    def _wait_for_rate_limit(self) -> None:
        elapsed = time.monotonic() - self._last_request_at
        if elapsed < self.request_delay:
            time.sleep(self.request_delay - elapsed)

    @staticmethod
    def _decode_body(raw: bytes) -> str:
        return raw.decode("utf-8", errors="replace")

    @staticmethod
    def _error_message(body: str) -> str:
        try:
            payload = json.loads(body)
        except (json.JSONDecodeError, TypeError):
            return body.strip() or "пустой ответ"
        if isinstance(payload, dict):
            return str(payload.get("message") or payload.get("description") or payload)
        return str(payload)

    def request(
        self,
        method: str,
        path: str,
        *,
        query: dict[str, Any] | None = None,
        payload: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        url = f"{self.base_url}{path}"
        if query:
            url += "?" + urllib.parse.urlencode(query)
        data = None
        headers = {
            "Authorization": f"OAuth {self.token}",
            "Accept": "application/json",
            "User-Agent": "yandex360-xlsx-import/1.0",
        }
        if payload is not None:
            data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            headers["Content-Type"] = "application/json; charset=utf-8"

        last_error: Exception | None = None
        for attempt in range(1, self.max_retries + 1):
            self._wait_for_rate_limit()
            request = urllib.request.Request(url, data=data, headers=headers, method=method)
            LOGGER.debug("%s %s", method, url)
            try:
                with urllib.request.urlopen(request, timeout=self.timeout) as response:
                    self._last_request_at = time.monotonic()
                    body = self._decode_body(response.read())
                    request_id = response.headers.get("x-request-id", "")
                    LOGGER.debug("HTTP %s, x-request-id=%s", response.status, request_id)
                    if not body.strip():
                        return {}
                    parsed = json.loads(body)
                    if not isinstance(parsed, dict):
                        raise ApiError("API вернул JSON неожиданного типа", request_id=request_id)
                    return parsed
            except urllib.error.HTTPError as exc:
                self._last_request_at = time.monotonic()
                body = self._decode_body(exc.read())
                request_id = exc.headers.get("x-request-id", "") if exc.headers else ""
                message = self._error_message(body)
                last_error = ApiError(
                    f"HTTP {exc.code}: {message}",
                    status=exc.code,
                    request_id=request_id,
                    response_body=body,
                )
                if exc.code not in RETRYABLE_HTTP_CODES or attempt == self.max_retries:
                    raise last_error
                retry_after = exc.headers.get("Retry-After") if exc.headers else None
                try:
                    delay = float(retry_after) if retry_after else 0.0
                except ValueError:
                    delay = 0.0
                if delay <= 0:
                    delay = min(30.0, 2 ** (attempt - 1)) + random.uniform(0.0, 0.5)
                LOGGER.warning(
                    "HTTP %s. Повтор %s/%s через %.1f с. x-request-id=%s",
                    exc.code,
                    attempt + 1,
                    self.max_retries,
                    delay,
                    request_id,
                )
                time.sleep(delay)
            except (urllib.error.URLError, socket.timeout, TimeoutError) as exc:
                self._last_request_at = time.monotonic()
                last_error = ApiError(f"Сетевая ошибка: {exc}")
                if attempt == self.max_retries:
                    raise last_error
                delay = min(30.0, 2 ** (attempt - 1)) + random.uniform(0.0, 0.5)
                LOGGER.warning(
                    "Сетевая ошибка. Повтор %s/%s через %.1f с: %s",
                    attempt + 1,
                    self.max_retries,
                    delay,
                    exc,
                )
                time.sleep(delay)
            except json.JSONDecodeError as exc:
                raise ApiError(f"API вернул некорректный JSON: {exc}") from exc
        raise ApiError(str(last_error or "Неизвестная ошибка API"))

    def list_departments(self) -> list[dict[str, Any]]:
        departments: list[dict[str, Any]] = []
        page = 1
        while True:
            response = self.request(
                "GET",
                f"/directory/v1/org/{self.org_id}/departments",
                query={"page": page, "perPage": 100, "orderBy": "id"},
            )
            page_items = response.get("departments", [])
            if not isinstance(page_items, list):
                raise ApiError("В ответе списка подразделений отсутствует массив departments")
            departments.extend(item for item in page_items if isinstance(item, dict))
            pages = int(response.get("pages") or page)
            if page >= pages:
                return departments
            page += 1

    def list_users(self) -> list[dict[str, Any]]:
        users: list[dict[str, Any]] = []
        page = 1
        while True:
            response = self.request(
                "GET",
                f"/directory/v1/org/{self.org_id}/users",
                query={"page": page, "perPage": 1000},
            )
            page_items = response.get("users", [])
            if not isinstance(page_items, list):
                raise ApiError("В ответе списка сотрудников отсутствует массив users")
            users.extend(item for item in page_items if isinstance(item, dict))
            pages = int(response.get("pages") or page)
            if page >= pages:
                return users
            page += 1

    def create_department(self, name: str, parent_id: int) -> dict[str, Any]:
        return self.request(
            "POST",
            f"/directory/v1/org/{self.org_id}/departments",
            payload={"name": name, "parentId": parent_id},
        )

    def create_user(self, payload: dict[str, Any]) -> dict[str, Any]:
        return self.request(
            "POST",
            f"/directory/v1/org/{self.org_id}/users",
            payload=payload,
        )


def department_key(parent_id: int, name: str) -> tuple[int, str]:
    return parent_id, name.strip().casefold()


def build_department_index(
    departments: list[dict[str, Any]],
) -> dict[tuple[int, str], dict[str, Any]]:
    index: dict[tuple[int, str], dict[str, Any]] = {}
    for department in departments:
        try:
            department_id = int(department["id"])
            parent_id = int(department["parentId"])
            name = clean_text(department["name"])
        except (KeyError, TypeError, ValueError) as exc:
            raise ApiError(f"Некорректное подразделение в ответе API: {department}") from exc
        key = department_key(parent_id, name)
        if key in index and int(index[key]["id"]) != department_id:
            raise ApiError(
                f"В API найдены два подразделения '{name}' с одним parentId={parent_id}"
            )
        index[key] = {**department, "id": department_id, "parentId": parent_id, "name": name}
    return index


def resolve_or_create_department_path(
    path: tuple[str, ...],
    *,
    client: ApiClient,
    index: dict[tuple[int, str], dict[str, Any]],
    dry_run: bool,
    synthetic_id: list[int],
    summary: Summary,
) -> int:
    parent_id = ROOT_DEPARTMENT_ID
    for name in path:
        key = department_key(parent_id, name)
        existing = index.get(key)
        if existing is not None:
            parent_id = int(existing["id"])
            continue
        if dry_run:
            new_id = synthetic_id[0]
            synthetic_id[0] -= 1
            created = {"id": new_id, "parentId": parent_id, "name": name, "planned": True}
            summary.departments_planned += 1
            LOGGER.info("[DRY-RUN] Будет создано подразделение: %s (parentId=%s)", name, parent_id)
        else:
            created = client.create_department(name, parent_id)
            if "id" not in created:
                refreshed = build_department_index(client.list_departments())
                index.update(refreshed)
                created = index.get(key, {})
            if "id" not in created:
                raise ApiError(
                    f"Подразделение '{name}' создано, но его id не найден в ответе и повторном списке"
                )
            created = {
                **created,
                "id": int(created["id"]),
                "parentId": parent_id,
                "name": clean_text(created.get("name") or name),
            }
            summary.departments_created += 1
            LOGGER.info("Создано подразделение: %s (id=%s, parentId=%s)", name, created["id"], parent_id)
        index[key] = created
        parent_id = int(created["id"])
    return parent_id


def user_payload(row: UserRow, department_id: int) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "nickname": row.login,
        "departmentId": department_id,
        "name": {
            "first": row.first_name,
            "last": row.last_name,
            "middle": row.middle_name,
        },
    }
    optional_strings = {
        "password": row.password,
        "position": row.position,
        "gender": row.gender,
        "birthday": row.birthday,
        "language": row.language,
    }
    payload.update({key: value for key, value in optional_strings.items() if value})
    if row.must_change_password is not None:
        payload["passwordChangeRequired"] = row.must_change_password
    contacts: list[dict[str, str]] = []
    if row.work_phone:
        contacts.append({"type": "phone", "value": row.work_phone, "label": "Work"})
    if row.mobile_phone:
        contacts.append({"type": "phone", "value": row.mobile_phone, "label": "Mobile"})
    if row.personal_email:
        contacts.append({"type": "email", "value": row.personal_email, "label": "Personal"})
    if contacts:
        payload["contacts"] = contacts
    return payload


def full_name(row: UserRow) -> str:
    return " ".join(part for part in (row.last_name, row.first_name, row.middle_name) if part)


def run_import(
    rows: list[UserRow],
    *,
    client: ApiClient,
    dry_run: bool,
) -> Summary:
    summary = Summary()
    departments = client.list_departments()
    department_index = build_department_index(departments)
    summary.departments_existing = len(departments)
    existing_users = client.list_users()
    existing_logins = {
        clean_text(item.get("nickname")).casefold(): item
        for item in existing_users
        if clean_text(item.get("nickname"))
    }
    synthetic_id = [-1]

    for row in rows:
        path_text = " / ".join(row.department_path) or "Корневое подразделение"
        report = ReportRow(
            row_number=row.row_number,
            login=row.login,
            full_name=full_name(row),
            department_path=path_text,
            status="",
        )
        existing_user = existing_logins.get(row.login.casefold())
        if existing_user is not None:
            report.status = "skipped"
            report.user_id = clean_text(existing_user.get("id"))
            report.message = "Сотрудник с таким логином уже существует"
            summary.users_skipped += 1
            summary.reports.append(report)
            LOGGER.warning("Строка %s: логин %s уже существует — пропуск", row.row_number, row.login)
            continue
        try:
            department_id = resolve_or_create_department_path(
                row.department_path,
                client=client,
                index=department_index,
                dry_run=dry_run,
                synthetic_id=synthetic_id,
                summary=summary,
            )
            payload = user_payload(row, department_id)
            if dry_run:
                report.status = "planned"
                report.message = "Будет создан"
                summary.users_planned += 1
                LOGGER.info(
                    "[DRY-RUN] Строка %s: будет создан пользователь %s в '%s'",
                    row.row_number,
                    row.login,
                    path_text,
                )
            else:
                created = client.create_user(payload)
                report.status = "created"
                report.user_id = clean_text(created.get("id"))
                report.message = "Создан"
                summary.users_created += 1
                existing_logins[row.login.casefold()] = created
                LOGGER.info(
                    "Строка %s: создан пользователь %s (id=%s)",
                    row.row_number,
                    row.login,
                    report.user_id or "не указан API",
                )
        except ApiError as exc:
            report.status = "error"
            report.message = str(exc)
            report.request_id = exc.request_id
            summary.users_failed += 1
            LOGGER.error(
                "Строка %s, логин %s: %s. x-request-id=%s",
                row.row_number,
                row.login,
                exc,
                exc.request_id,
            )
        summary.reports.append(report)
    return summary


def default_report_path(xlsx: Path, dry_run: bool) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    mode = "dry_run" if dry_run else "import"
    return xlsx.resolve().parent / f"yandex360_{mode}_report_{timestamp}.csv"


def write_report(path: Path, reports: list[ReportRow]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file, delimiter=";")
        writer.writerow(
            [
                "row_number",
                "login",
                "full_name",
                "department_path",
                "status",
                "user_id",
                "message",
                "x_request_id",
            ]
        )
        for item in reports:
            writer.writerow(
                [
                    item.row_number,
                    item.login,
                    item.full_name,
                    item.department_path,
                    item.status,
                    item.user_id,
                    item.message,
                    item.request_id,
                ]
            )


def print_summary(summary: Summary, report_path: Path, dry_run: bool) -> None:
    print("\nИтог:")
    print(f"  Подразделений найдено в организации: {summary.departments_existing}")
    if dry_run:
        print(f"  Подразделений будет создано: {summary.departments_planned}")
        print(f"  Сотрудников будет создано: {summary.users_planned}")
    else:
        print(f"  Подразделений создано: {summary.departments_created}")
        print(f"  Сотрудников создано: {summary.users_created}")
    print(f"  Сотрудников пропущено: {summary.users_skipped}")
    print(f"  Ошибок: {summary.users_failed}")
    print(f"  CSV-отчёт: {report_path}")


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    configure_logging(args.verbose)
    try:
        rows = read_users_from_xlsx(args.xlsx, args.sheet)
        LOGGER.info("XLSX проверен: %s строк с пользователями", len(rows))
        token, org_id = resolve_credentials(args)
        client = ApiClient(
            token,
            org_id,
            request_delay=args.request_delay,
            max_retries=args.max_retries,
            timeout=args.timeout,
        )
        summary = run_import(rows, client=client, dry_run=args.dry_run)
        report_path = args.report or default_report_path(args.xlsx, args.dry_run)
        write_report(report_path, summary.reports)
        print_summary(summary, report_path, args.dry_run)
        return 1 if summary.users_failed else 0
    except (ImportValidationError, ApiError) as exc:
        LOGGER.error("%s", exc)
        if isinstance(exc, ApiError) and exc.request_id:
            LOGGER.error("x-request-id: %s", exc.request_id)
        return 2
    except KeyboardInterrupt:
        LOGGER.error("Операция прервана пользователем")
        return 130


if __name__ == "__main__":
    raise SystemExit(main())
